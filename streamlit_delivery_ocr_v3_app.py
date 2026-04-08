import os
import re
import tempfile
from io import BytesIO

import pandas as pd
import pytesseract
import streamlit as st
from PIL import Image, ImageOps, ImageFilter
from pdf2image import convert_from_path
from openpyxl import load_workbook


st.set_page_config(page_title="PDF OCR Tracker Updater", layout="wide")
st.title("PDF OCR Tracker Updater")


# -----------------------------
# Session State
# -----------------------------
if "processed" not in st.session_state:
    st.session_state.processed = False

if "results" not in st.session_state:
    st.session_state.results = {}

if "run_id" not in st.session_state:
    st.session_state.run_id = 0


# -----------------------------
# Helpers
# -----------------------------
def normalize_text(value):
    if value is None:
        return ""
    value = str(value).strip()
    if value.startswith("'"):
        value = value[1:]
    value = value.replace("\xa0", " ")
    value = " ".join(value.split())
    return value.upper()


def normalize_header(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def normalize_part_number(value):
    if value is None:
        return ""
    value = str(value).strip()
    if value.startswith("'"):
        value = value[1:]
    value = value.replace("\xa0", " ")
    value = " ".join(value.split())
    return value.upper()


def combine_unique_values(series):
    values = []
    for v in series:
        if pd.isna(v):
            continue
        v = str(v).strip()
        if v != "":
            values.append(v)

    unique_values = sorted(set(values), key=lambda x: (len(x), x))
    return ", ".join(unique_values)


def merge_comma_separated(existing_value, new_value):
    existing_items = []
    new_items = []

    if existing_value not in [None, ""]:
        existing_items = [x.strip() for x in str(existing_value).split(",") if x.strip()]

    if new_value not in [None, ""]:
        new_items = [x.strip() for x in str(new_value).split(",") if x.strip()]

    merged = sorted(set(existing_items + new_items), key=lambda x: (len(x), x))
    return ", ".join(merged)


def normalize_quantity_text(qty_raw):
    qty_raw = str(qty_raw).strip()

    if "," in qty_raw and "." in qty_raw:
        if qty_raw.rfind(",") > qty_raw.rfind("."):
            qty_clean = qty_raw.replace(".", "").replace(",", ".")
        else:
            qty_clean = qty_raw.replace(",", "")
    elif "," in qty_raw:
        parts = qty_raw.split(",")
        if len(parts[-1]) == 3 and len(parts) > 1:
            qty_clean = qty_raw.replace(",", "")
        else:
            qty_clean = qty_raw.replace(",", ".")
    elif "." in qty_raw:
        parts = qty_raw.split(".")
        if len(parts[-1]) == 3 and len(parts) > 1:
            qty_clean = qty_raw.replace(".", "")
        else:
            qty_clean = qty_raw
    else:
        qty_clean = qty_raw

    return int(float(qty_clean))


def extract_document_number(text):
    patterns = [
        r"Truck\s*No\.?\s*[:#]?\s*([A-Z0-9\-\/]+)",
        r"Truck\s*[:#]?\s*([A-Z0-9\-\/]+)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()

    return ""


def is_table_header(line):
    line_lower = line.lower()

    item_words = [
        "item",
        "artikelnr",
        "artikelnr.",
    ]

    desc_words = [
        "descr",
        "beschr",
        "beschr.",
    ]

    has_item = any(word in line_lower for word in item_words)
    has_desc = any(word in line_lower for word in desc_words)

    return has_item and has_desc


def is_end_of_table(line):
    line_lower = line.lower().strip()

    end_words = [
        "total colli",
        "anzahl colli",
        "anzahl colli:",
    ]

    return any(word in line_lower for word in end_words)


def extract_colli_number(line):
    patterns = [
        r"colli\s*#?\s*([0-9]+)",
        r"colli\s*nr\.?\s*[:#]?\s*([0-9]+)",
    ]

    for pattern in patterns:
        match = re.search(pattern, line, re.IGNORECASE)
        if match:
            return match.group(1).strip()

    return ""


def preprocess_for_ocr(img):
    img = img.convert("L")
    img = ImageOps.autocontrast(img)
    img = img.filter(ImageFilter.SHARPEN)
    img = img.point(lambda p: 255 if p > 170 else 0)
    return img


def line_looks_like_item_row(line):
    return re.search(r"\b\d{4}\s*[-–—]\s*\d{4}\b", line) is not None


def extract_item_and_remainder(line):
    """
    More forgiving item-number extraction:
    - allows -, en dash, em dash
    - allows leading text/noise before the part number
    - does not require exact spacing after item number
    """
    item_match = re.search(r"(\d{4}\s*[-–—]\s*\d{4})", line)
    if not item_match:
        return None, None

    item_no = re.sub(r"\s*[-–—]\s*", "-", item_match.group(1).strip())
    remainder = line[item_match.end():].strip()
    return item_no, remainder


def parse_quantity_and_description(remainder):
    """
    Try a few increasingly forgiving patterns to recover description + quantity.
    """

    qty_match = re.search(
        r"(.+?)\s+([\d.,]+)\s+(piece|stück|pcs?|bundle|kg|pc|roll|rolle|set|sets|meter|metre|each|einh\.?|stk|stck|m)\b",
        remainder,
        re.IGNORECASE,
    )
    if qty_match:
        description = qty_match.group(1).strip()
        qty_raw = qty_match.group(2).strip()
        unit = qty_match.group(3).strip().lower()
        try:
            quantity = normalize_quantity_text(qty_raw)
            return description, quantity, unit
        except Exception:
            pass

    fallback_match = re.search(
        r"(.+?)\s+([\d.,]+)(?:\s+\S+)?$",
        remainder,
        re.IGNORECASE,
    )
    if fallback_match:
        description = fallback_match.group(1).strip()
        qty_raw = fallback_match.group(2).strip()
        unit = "auto"
        try:
            quantity = normalize_quantity_text(qty_raw)
            return description, quantity, unit
        except Exception:
            pass

    fallback_match_2 = re.search(
        r"(.+?)\s+([\d.,]+)(?:\s+[A-Za-zÄÖÜäöü\.]+)?$",
        remainder,
        re.IGNORECASE,
    )
    if fallback_match_2:
        description = fallback_match_2.group(1).strip()
        qty_raw = fallback_match_2.group(2).strip()
        unit = "fallback"
        try:
            quantity = normalize_quantity_text(qty_raw)
            return description, quantity, unit
        except Exception:
            pass

    return None, None, None


def clean_ocr_line(line):
    if line is None:
        return ""

    line = str(line).replace("\xa0", " ")
    line = line.replace("|", " ")
    line = line.replace("—", "-").replace("–", "-")
    line = re.sub(r"[ \t]+", " ", line)
    line = re.sub(r"\s*-\s*", "-", line)
    return line.strip()


def classify_ocr_line(line):
    line_lower = line.lower()

    if not line:
        return "blank"
    if is_table_header(line):
        return "table_header"
    if is_end_of_table(line):
        return "end_of_table"
    if extract_colli_number(line):
        return "colli_line"
    if line_looks_like_item_row(line):
        return "possible_item_row"
    if "truck" in line_lower:
        return "document_line"
    return "other"


def load_pages_from_upload(uploaded_file):
    ext = os.path.splitext(uploaded_file.name)[1].lower()

    if ext == ".pdf":
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name

        try:
            pages = convert_from_path(tmp_path, dpi=300)
            return [page.convert("L") for page in pages]
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
    else:
        img = Image.open(uploaded_file).convert("L")
        return [img]


def process_delivery_files(delivery_files):
    all_items = []
    preview_images = []
    ocr_text_records = []
    cleaned_ocr_records = []
    classified_line_records = []
    skipped_line_records = []

    progress = st.progress(0, text="Starting OCR processing...")
    total_files = len(delivery_files)

    for file_index, uploaded_file in enumerate(delivery_files, start=1):
        progress.progress(
            int(((file_index - 1) / total_files) * 100),
            text=f"Processing file {file_index} of {total_files}: {uploaded_file.name}",
        )

        pages = load_pages_from_upload(uploaded_file)

        table_started_in_previous_page = False

        for page_index, page_img in enumerate(pages, start=1):
            processed_img = preprocess_for_ocr(page_img)
            preview_images.append((f"{uploaded_file.name} - Page {page_index}", processed_img))

            text = pytesseract.image_to_string(
                processed_img,
                config="--psm 6 -c preserve_interword_spaces=1"
            )

            ocr_text_records.append(
                {
                    "SourceFile": uploaded_file.name,
                    "PageNumber": page_index,
                    "OCR_Text": text,
                }
            )

            document_number = extract_document_number(text)

            lines = text.split("\n")
            capture = table_started_in_previous_page
            current_colli = ""
            page_item_count = 0

            for line_number, line in enumerate(lines, start=1):
                original_line = line
                cleaned_line = clean_ocr_line(line)

                classified_line_records.append(
                    {
                        "SourceFile": uploaded_file.name,
                        "PageNumber": page_index,
                        "LineNumber": line_number,
                        "LineType": classify_ocr_line(cleaned_line),
                        "RawLine": original_line,
                        "CleanedLine": cleaned_line,
                    }
                )

                if cleaned_line:
                    cleaned_ocr_records.append(
                        {
                            "SourceFile": uploaded_file.name,
                            "PageNumber": page_index,
                            "LineNumber": line_number,
                            "RawLine": original_line,
                            "CleanedLine": cleaned_line,
                        }
                    )

                line = cleaned_line

                if not line:
                    continue

                detected_colli = extract_colli_number(line)
                if detected_colli:
                    current_colli = detected_colli

                if is_table_header(line):
                    capture = True
                    continue

                if not capture and line_looks_like_item_row(line):
                    capture = True

                if not capture:
                    continue

                if is_end_of_table(line):
                    capture = False
                    continue

                line = re.split(
                    r"(total\s+colli|anzahl\s+colli\s*:?)",
                    line,
                    flags=re.IGNORECASE
                )[0].strip()

                if not line:
                    continue

                item_no, remainder = extract_item_and_remainder(line)
                if not item_no:
                    if re.search(r"\d{4}", line):
                        skipped_line_records.append(
                            {
                                "SourceFile": uploaded_file.name,
                                "PageNumber": page_index,
                                "LineNumber": line_number,
                                "Reason": "No item match",
                                "Line": original_line,
                                "CleanedLine": line,
                            }
                        )
                    continue

                description, quantity, unit = parse_quantity_and_description(remainder)

                if quantity is None or description is None:
                    skipped_line_records.append(
                        {
                            "SourceFile": uploaded_file.name,
                            "PageNumber": page_index,
                            "LineNumber": line_number,
                            "Reason": "No quantity/description parse",
                            "Line": original_line,
                            "CleanedLine": line,
                            "ItemNoGuess": item_no,
                            "Remainder": remainder,
                        }
                    )
                    continue

                all_items.append(
                    {
                        "ItemNo": item_no,
                        "Description": description,
                        "Quantity": quantity,
                        "Unit": unit,
                        "ColliNo": current_colli,
                        "DocumentNumber": document_number,
                        "SourceFile": uploaded_file.name,
                        "PageNumber": page_index,
                        "LineNumber": line_number,
                    }
                )

                page_item_count += 1

            table_started_in_previous_page = page_item_count > 0

    progress.progress(100, text="OCR processing complete.")

    raw_df = pd.DataFrame(all_items)
    ocr_text_df = pd.DataFrame(ocr_text_records)
    cleaned_ocr_df = pd.DataFrame(cleaned_ocr_records)
    classified_lines_df = pd.DataFrame(classified_line_records)
    skipped_lines_df = pd.DataFrame(skipped_line_records)

    if raw_df.empty:
        summary_df = pd.DataFrame(columns=["ItemNo", "Description", "Quantity", "PalletList", "DocumentList"])
    else:
        summary_df = (
            raw_df.groupby("ItemNo")
            .agg(
                Quantity=("Quantity", "sum"),
                Description=("Description", lambda x: x.mode().iat[0] if not x.mode().empty else x.iloc[0]),
                PalletList=("ColliNo", combine_unique_values),
                DocumentList=("DocumentNumber", combine_unique_values),
            )
            .reset_index()
        )

        summary_df = summary_df[["ItemNo", "Description", "Quantity", "PalletList", "DocumentList"]]

    return raw_df, summary_df, preview_images, ocr_text_df, cleaned_ocr_df, classified_lines_df, skipped_lines_df


def build_tracker_lookup(wb):
    tracker_rows = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = {}

        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is not None:
                headers[normalize_header(value)] = col

        item_col = headers.get("item #") or headers.get("part #")
        qty_col = headers.get("qty received")
        pallet_col = headers.get("pallet #")
        container_col = headers.get("container")

        if not item_col or not qty_col:
            continue

        for row in range(2, ws.max_row + 1):
            item_value = ws.cell(row=row, column=item_col).value
            if item_value is None:
                continue

            normalized_item = normalize_text(item_value)

            if normalized_item:
                if normalized_item not in tracker_rows:
                    tracker_rows[normalized_item] = []

                tracker_rows[normalized_item].append(
                    {
                        "sheet": ws,
                        "row": row,
                        "headers": headers,
                        "item_col": item_col,
                        "qty_col": qty_col,
                        "pallet_col": pallet_col,
                        "container_col": container_col,
                        "original_item": item_value,
                    }
                )

    return tracker_rows


def update_tracker_workbook(wb, summary_df, raw_df):
    tracker_rows = build_tracker_lookup(wb)

    matched = []
    not_found = []

    for _, row in summary_df.iterrows():
        raw_item_no = row["ItemNo"]
        item_no = normalize_part_number(raw_item_no)

        qty = row["Quantity"]
        desc = row["Description"]
        pallet_list = row["PalletList"]
        document_list = row["DocumentList"]

        if item_no in tracker_rows:
            entries = tracker_rows[item_no]

            for entry in entries:
                ws = entry["sheet"]
                excel_row = entry["row"]

                qty_col = entry["qty_col"]
                pallet_col = entry["pallet_col"]
                container_col = entry["container_col"]

                ws.cell(row=excel_row, column=qty_col).value = qty

                if pallet_col is not None:
                    existing_pallets = ws.cell(row=excel_row, column=pallet_col).value
                    merged_pallets = merge_comma_separated(existing_pallets, pallet_list)
                    ws.cell(row=excel_row, column=pallet_col).value = merged_pallets

                if container_col is not None:
                    existing_containers = ws.cell(row=excel_row, column=container_col).value
                    merged_containers = merge_comma_separated(existing_containers, document_list)
                    ws.cell(row=excel_row, column=container_col).value = merged_containers

                matched.append(
                    {
                        "ItemNo": item_no,
                        "Description": desc,
                        "Quantity": qty,
                        "PalletList": pallet_list,
                        "DocumentList": document_list,
                        "Sheet": ws.title,
                        "Row": excel_row,
                    }
                )
        else:
            not_found.append(
                {
                    "ItemNo": item_no,
                    "Description": desc,
                    "Quantity": qty,
                    "PalletList": pallet_list,
                    "DocumentList": document_list,
                }
            )

    unmatched_rows = []

    item_column = "Item #" if "Item #" in raw_df.columns else "ItemNo" if "ItemNo" in raw_df.columns else None
    if item_column is None:
        raise ValueError(f"Could not find item column in raw_df. Columns found: {list(raw_df.columns)}")

    for _, row in raw_df.iterrows():
        raw_item_no = row[item_column]
        item_no = normalize_part_number(raw_item_no)

        desc = row["Description"] if "Description" in raw_df.columns else ""
        qty = row["Quantity"] if "Quantity" in raw_df.columns else ""
        source_file = row["SourceFile"] if "SourceFile" in raw_df.columns else ""
        page_number = row["PageNumber"] if "PageNumber" in raw_df.columns else ""
        line_number = row["LineNumber"] if "LineNumber" in raw_df.columns else ""
        colli_no = row["ColliNo"] if "ColliNo" in raw_df.columns else ""
        document_number = row["DocumentNumber"] if "DocumentNumber" in raw_df.columns else ""

        if item_no not in tracker_rows:
            unmatched_rows.append(
                {
                    "Item #": raw_item_no,
                    "Normalized Item #": item_no,
                    "Description": desc,
                    "Quantity": qty,
                    "ColliNo": colli_no,
                    "DocumentNumber": document_number,
                    "SourceFile": source_file,
                    "PageNumber": page_number,
                    "LineNumber": line_number,
                }
            )

    unmatched_df = pd.DataFrame(unmatched_rows)

    if not unmatched_df.empty:
        unmatched_df = (
            unmatched_df.groupby(
                ["Item #", "Normalized Item #", "Description", "ColliNo", "DocumentNumber", "SourceFile", "PageNumber", "LineNumber"],
                as_index=False,
            )["Quantity"].sum()
        )

    matched_df = pd.DataFrame(matched)
    not_found_df = pd.DataFrame(not_found)

    return wb, matched_df, not_found_df, unmatched_df


def workbook_to_bytes(wb):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def dataframe_to_excel_bytes(sheets_dict):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    output.seek(0)
    return output.getvalue()


# -----------------------------
# UI
# -----------------------------
delivery_files = st.file_uploader(
    "Upload delivery PDFs/images",
    type=["pdf", "png", "jpg", "jpeg", "tif", "tiff"],
    accept_multiple_files=True,
    key=f"delivery_files_{st.session_state.run_id}",
)

tracker_file = st.file_uploader(
    "Upload tracker workbook",
    type=["xlsx", "xlsm"],
    key=f"tracker_file_{st.session_state.run_id}",
)

col_a, col_b = st.columns([1, 1])

with col_a:
    process_clicked = st.button("Process Files", type="primary")

with col_b:
    reset_clicked = st.button("Reset / Run New Files")


if reset_clicked:
    st.session_state.processed = False
    st.session_state.results = {}
    st.session_state.run_id += 1
    st.rerun()


if process_clicked:
    if not delivery_files:
        st.error("Please upload at least one delivery PDF/image.")
        st.stop()

    if not tracker_file:
        st.error("Please upload the tracker workbook.")
        st.stop()

    st.session_state.processed = False
    st.session_state.results = {}

    with st.spinner("Running OCR and updating tracker..."):
        raw_df, summary_df, preview_images, ocr_text_df, cleaned_ocr_df, classified_lines_df, skipped_lines_df = process_delivery_files(delivery_files)

        tracker_bytes = tracker_file.getvalue()
        wb = load_workbook(BytesIO(tracker_bytes))

        wb, matched_df, not_found_df, unmatched_df = update_tracker_workbook(wb, summary_df, raw_df)

        ocr_results_bytes = dataframe_to_excel_bytes(
            {
                "Parsed_OCR_Rows": raw_df if not raw_df.empty else pd.DataFrame(),
                "Summarized_Totals": summary_df if not summary_df.empty else pd.DataFrame(),
            }
        )

        parsed_rows_export_bytes = dataframe_to_excel_bytes(
            {
                "Parsed_OCR_Rows": raw_df if not raw_df.empty else pd.DataFrame()
            }
        )

        ocr_text_export_bytes = dataframe_to_excel_bytes(
            {
                "OCR_Raw_Text": ocr_text_df if not ocr_text_df.empty else pd.DataFrame(),
                "OCR_Cleaned_Lines": cleaned_ocr_df if not cleaned_ocr_df.empty else pd.DataFrame(),
                "OCR_Classified_Lines": classified_lines_df if not classified_lines_df.empty else pd.DataFrame(),
            }
        )

        skipped_lines_export_bytes = dataframe_to_excel_bytes(
            {
                "Skipped_Lines": skipped_lines_df if not skipped_lines_df.empty else pd.DataFrame()
            }
        )

        updated_tracker_bytes = workbook_to_bytes(wb)

        unmatched_export_bytes = None
        if unmatched_df is not None and not unmatched_df.empty:
            unmatched_export_bytes = dataframe_to_excel_bytes({"Unmatched_OCR_Items": unmatched_df})

    st.session_state.results = {
        "raw_df": raw_df,
        "summary_df": summary_df,
        "matched_df": matched_df,
        "not_found_df": not_found_df,
        "unmatched_df": unmatched_df,
        "preview_images": preview_images,
        "updated_tracker_bytes": updated_tracker_bytes,
        "ocr_results_bytes": ocr_results_bytes,
        "parsed_rows_export_bytes": parsed_rows_export_bytes,
        "ocr_text_export_bytes": ocr_text_export_bytes,
        "ocr_text_df": ocr_text_df,
        "cleaned_ocr_df": cleaned_ocr_df,
        "classified_lines_df": classified_lines_df,
        "skipped_lines_df": skipped_lines_df,
        "skipped_lines_export_bytes": skipped_lines_export_bytes,
        "unmatched_export_bytes": unmatched_export_bytes,
    }

    st.session_state.processed = True
    st.rerun()


if st.session_state.processed:
    results = st.session_state.results

    raw_df = results["raw_df"]
    summary_df = results["summary_df"]
    matched_df = results["matched_df"]
    not_found_df = results["not_found_df"]
    unmatched_df = results["unmatched_df"]
    preview_images = results["preview_images"]
    updated_tracker_bytes = results["updated_tracker_bytes"]
    ocr_results_bytes = results["ocr_results_bytes"]
    parsed_rows_export_bytes = results["parsed_rows_export_bytes"]
    ocr_text_export_bytes = results["ocr_text_export_bytes"]
    ocr_text_df = results["ocr_text_df"]
    cleaned_ocr_df = results["cleaned_ocr_df"]
    classified_lines_df = results["classified_lines_df"]
    skipped_lines_df = results["skipped_lines_df"]
    skipped_lines_export_bytes = results["skipped_lines_export_bytes"]
    unmatched_export_bytes = results["unmatched_export_bytes"]

    st.success("Processing complete.")

    if preview_images:
        with st.expander("Preview OCR Images", expanded=False):
            for name, img in preview_images:
                st.markdown(f"**{name}**")
                st.image(img, use_container_width=True)

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Summarized Totals")
        st.dataframe(summary_df, use_container_width=True)

    with col2:
        st.subheader("Matched Rows")
        st.dataframe(matched_df, use_container_width=True)

    with st.expander("Parsed OCR Rows Preview", expanded=False):
        st.dataframe(raw_df, use_container_width=True)

    with st.expander("OCR Raw Text Preview", expanded=False):
        st.dataframe(ocr_text_df, use_container_width=True)

    with st.expander("Cleaned OCR Lines Preview", expanded=False):
        st.dataframe(cleaned_ocr_df, use_container_width=True)

    with st.expander("Classified OCR Lines Preview", expanded=False):
        st.dataframe(classified_lines_df, use_container_width=True)

    with st.expander("Skipped Lines Preview", expanded=False):
        if not skipped_lines_df.empty:
            st.dataframe(skipped_lines_df, use_container_width=True)
        else:
            st.info("No skipped lines were recorded.")

    st.subheader("Unmatched Items")
    if unmatched_df is not None and not unmatched_df.empty:
        st.dataframe(unmatched_df, use_container_width=True)
    else:
        st.info("No unmatched items.")

    if not not_found_df.empty:
        with st.expander("Summary items not found in tracker", expanded=False):
            st.dataframe(not_found_df, use_container_width=True)

    st.download_button(
        label="Download Updated Tracker",
        data=updated_tracker_bytes,
        file_name="Updated_Tracking_File.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        label="Download OCR Results",
        data=ocr_results_bytes,
        file_name="OCR_Results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        label="Download Parsed OCR Rows",
        data=parsed_rows_export_bytes,
        file_name="Parsed_OCR_Rows.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        label="Download OCR Raw Text",
        data=ocr_text_export_bytes,
        file_name="OCR_Raw_Text.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        label="Download Skipped Lines",
        data=skipped_lines_export_bytes,
        file_name="Skipped_Lines.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    if unmatched_export_bytes is not None:
        st.download_button(
            label="Download Unmatched Items",
            data=unmatched_export_bytes,
            file_name="Unmatched_OCR_Items.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("Upload your files and click Process Files. The app will only rerun processing when you press that button or reset the session.")
